import random
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference, BarChart, RadarChart

def generate_random_data(size=5):
    return [random.randint(1, 31) for _ in range(size)]

def create_line_chart(ws, x_data, y_data, x_label="", y_label="", title=""):
    for row, (x, y) in enumerate(zip(x_data, y_data), start=1):
        ws.cell(row=row, column=1, value=x)
        ws.cell(row=row, column=2, value=y)

    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(y_data))
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(x_data))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = title
    chart.x_axis.title = x_label
    chart.y_axis.title = y_label
    ws.add_chart(chart, "E4")


def create_bar_chart(ws, x_data, y_data, error_data, x_label="", y_label="", title=""):
    for i, x_value in enumerate(x_data, start=2):
        ws.cell(row=i, column=1, value=x_value)

    for i, y_value in enumerate(y_data, start=2):
        ws.cell(row=i, column=2, value=y_value)

    for i, error_value in enumerate(error_data, start=2):
        ws.cell(row=i, column=3, value=error_value)

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=len(y_data)+1), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=len(error_data)+1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(x_data)+1))
    chart.title = title
    chart.x_axis.title = x_label
    chart.y_axis.title = y_label
    ws.add_chart(chart, "E5")


def create_histogram(ws, data, n_bins, x_label="", y_label="", title=""):
    for i, value in enumerate(data, start=1):
        ws.cell(row=i, column=1, value=value)

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=len(data)))
    chart.bins = n_bins
    chart.title = title
    chart.x_axis.title = x_label
    chart.y_axis.title = y_label
    ws.add_chart(chart, "E4")


def create_radar_chart(ws):
    rows = [
        [ "Bulbs"],
        ['fait'] + generate_random_data(1),
        ['push'] + generate_random_data(1),
        ['damage'] + generate_random_data(1),
        ['personage'] + generate_random_data(1),
        ['hill'] + generate_random_data(1),
    ]

    for row in rows:
        ws.append(row)

    chart = RadarChart()
    chart.type = "filled"
    labels = Reference(ws, min_col=1, min_row=2, max_row=6)
    data = Reference(ws, min_col=2, min_row=1, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.style = 26
    chart.title = "Garden Centre Sales"
    chart.y_axis.delete = True 

    ws.add_chart(chart, "A17")


# Создаем книгу Excel
workbook = Workbook()

# Создаем листы
sheet_1 = workbook.active
sheet_1.title = "Sheet_1"

sheet_2 = workbook.create_sheet(title="Sheet_2")
sheet_3 = workbook.create_sheet(title="Sheet_3")
sheet_4 = workbook.create_sheet(title="Sheet_4")

# Пример использования
x_data = generate_random_data()
y_data = generate_random_data()
error_data = generate_random_data()
data_for_histogram = generate_random_data()

create_line_chart(sheet_1, x_data, y_data, "X Label", "Y Label", "Line Chart Title")
create_bar_chart(sheet_2, x_data, y_data, error_data, "X Label", "Y Label", "Bar Chart Title")
create_histogram(sheet_3, data_for_histogram, n_bins=5, x_label="X Label", y_label="Frequency", title="Histogram Title")
create_radar_chart(sheet_4)

# Сохраняем книгу Excel
workbook.save("charts_random.xlsx")
